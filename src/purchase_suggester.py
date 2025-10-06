import pandas as pd
from .stock_manager import calculate_min_stock
from .supplier_selector import select_suppliers
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def suggest_packaging(qty: float, product_id: str, packaging_df: pd.DataFrame):
    """
    Convierte la cantidad sugerida (en unidades) al número de empaques adecuados.
    Devuelve (cantidad_empaques, nombre_empaque, factor_empaque)
    """
    packs = packaging_df[packaging_df["product_id"] == product_id].copy()
    packs = packs[~packs["name"].str.contains("Neg", case=False, na=False)]

    #packs = packaging_df[packaging_df["product_id"] == product_id][["name","unit_conversion"]]
    if packs.empty or qty <= 0:
        return qty, "UND", 1

    # Ordenar empaques de menor a mayor
    packs = packs.sort_values("unit_conversion").reset_index(drop=True)

    for i, row in packs.iterrows():
        pack_size = row["unit_conversion"]
        pack_name = row["name"]

        if qty <= pack_size:
            return 1, pack_name, pack_size

        if i < len(packs) - 1:
            next_size = packs.loc[i+1,"unit_conversion"]
            if qty < next_size:
                factor = round(qty / pack_size)
                return factor, pack_name, pack_size

    # Si supera todos los empaques, usa el mayor
    last = packs.iloc[-1]
    factor = round(qty / last["unit_conversion"])
    return factor, last["name"], last["unit_conversion"]

def generate_purchase_suggestion(
    sales_df: pd.DataFrame,
    forecast_df: pd.DataFrame,
    stock_df: pd.DataFrame,
    products_df: pd.DataFrame,
    purchases_df: pd.DataFrame,
    packaging_df: pd.DataFrame,
    coverage_days: int = 7,
    safety_days: int = 3
) -> pd.DataFrame:

    # Normalizar IDs
    for df in (forecast_df, stock_df, products_df, purchases_df, packaging_df):
        if "product_id" in df.columns:
            df["product_id"] = df["product_id"].astype(str)

    # Asegurar columnas
    if "current_stock" not in stock_df.columns:
        stock_df = stock_df.rename(columns={stock_df.columns[1]: "current_stock"})

    # === 1) Calcular stock mínimo (tu función existente) ===
    min_stock_df = calculate_min_stock(sales_df, coverage_days=safety_days)

    # === 2) Expected demand ===
    forecast_df = forecast_df.copy()
    forecast_df["forecast"] = pd.to_numeric(forecast_df["forecast"], errors="coerce").fillna(0.0)
    forecast_df["expected_demand"] = forecast_df["forecast"] * coverage_days

    # === 3) Merge forecast + stock + min_stock ===
    merged = forecast_df.merge(stock_df[["product_id", "current_stock"]], on="product_id", how="left")
    merged = merged.merge(min_stock_df, on="product_id", how="left")

    merged["current_stock"] = merged["current_stock"].fillna(0)
    merged["min_stock"] = merged["min_stock"].fillna(0)

    # === 4) Suggested purchase (en unidades) ===
    merged["raw_suggested"] = (merged["expected_demand"] + merged["min_stock"]) - merged["current_stock"]
    merged["raw_suggested"] = np.where(merged["raw_suggested"] > 0, np.ceil(merged["raw_suggested"]), 0).astype(int)

    # === 5) Selección de proveedor ===
    suppliers_df = select_suppliers(purchases_df)
    suppliers_df["product_id"] = suppliers_df["product_id"].astype(str)
    merged = merged.merge(suppliers_df[["product_id", "best_supplier", "best_cost"]], on="product_id", how="left")

    # === 6) Merge con productos ===
    if "product_code" not in products_df.columns and "product_name" not in products_df.columns:
        products_df = products_df.rename(columns={
            products_df.columns[1]: "product_code",
            products_df.columns[2]: "product_name"
        })
    merged = merged.merge(products_df[["product_id", "product_code", "product_name"]], on="product_id", how="left")

    # === 7) Aplicar empaques ===
    pack_results = merged.apply(
        lambda row: suggest_packaging(row["raw_suggested"], row["product_id"], packaging_df),
        axis=1,
        result_type="expand"
    )
    merged[["suggested_purchase", "packaging", "pack_factor"]] = pack_results

    # === 8) Ajustar costo por empaque ===
    merged["best_cost"] = pd.to_numeric(merged["best_cost"], errors="coerce")
    merged["best_cost_per_pack"] = (merged["best_cost"] * merged["pack_factor"]).round(2)
    merged["est_total_cost"] = (merged["best_cost_per_pack"] * merged["suggested_purchase"]).round(2)

    # === 9) Redondeos y formato ===
    merged["best_cost_rounded"] = merged["best_cost_per_pack"].round(2)
    merged["min_stock"] = np.ceil(merged["min_stock"]).astype(int)
    merged["expected_demand"] = np.ceil(merged["expected_demand"]).astype(int)

    # === 10) DataFrame final ===
    merged["min_stock"] = np.ceil(merged["min_stock"]).astype(int)
    merged["expected_demand"] = np.ceil(merged["expected_demand"]).astype(int)
    merged["best_supplier"] = merged["best_supplier"].fillna("UNKNOWN").str.slice(0, 25)
    merged["total_units"] = merged["suggested_purchase"] * merged["pack_factor"]
    merged = merged[merged["total_units"] >= 3]

    final = merged[[
        "product_code", "product_name",
        "current_stock", "min_stock", "expected_demand",
        "suggested_purchase", "packaging",
        "best_supplier", "best_cost_per_pack", "est_total_cost"
    ]].rename(columns={"best_cost_per_pack": "best_cost"})

    merged["total_units"] = merged["suggested_purchase"] * merged["pack_factor"]
    merged = merged[merged["total_units"] >= 3]

    return final

def export_to_excel(df: pd.DataFrame, path: str):
    """Exporta el DataFrame ajustando automáticamente el ancho de columnas."""
    # Exportar a Excel
    df.to_excel(path, index=False)

    # Cargar libro y hoja
    wb = load_workbook(path)
    ws = wb.active

    # Ajustar ancho de columnas
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0

        # Obtener el nombre de la columna (A, B, C, ...)
        col_letter = get_column_letter(col_idx)

        # Considerar también el nombre de la columna (encabezado)
        column_name = ws[f"{col_letter}1"].value
        if column_name:
            max_length = len(str(column_name))

        # Revisar todas las celdas de la columna
        for cell in column_cells:
            if cell.value is not None:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass  # Por si hay valores extraños

        # Establecer ancho con un pequeño margen
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(path)
